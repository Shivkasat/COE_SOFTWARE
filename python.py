"""
© 2025 Shiv Sunil Kasat. All Rights Reserved.

This software and associated documentation files (the "Software") are proprietary to Shiv Sunil Kasat.
Unauthorized copying, distribution, modification, or usage of this Software, via any medium, is strictly prohibited.

Developed for Allan Smith Engineering as per agreement.

"""
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import os
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import numpy as np
import matplotlib.pyplot as plt
from fpdf import FPDF
import re
from PIL import Image as PILImage


app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

db = SQLAlchemy(app)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(256), nullable=False)


def radar_chart_report():
    print("Generating Radar Chart Report for Shell Profile & Temperature Analysis")

def gear_tyre_axial_runout():
    print("Generating Report for Gear & Tyre Axial Run-out Analysis")



def support_roller_deflection():
    print("Generating Report for Support Roller Deflection Analysis")

def support_roller_raceway():
    print("Generating Report for Support Roller Raceway Profile")

# Dictionary to map modules to functions

# Function to create a radar chart
def create_radar_chart(ax, AI, Shell_run_out, title):
    # Reverse and rotate data
    AI = AI[:-1]
    Shell_run_out = Shell_run_out[:-1]

    AI_flipped = AI[1:] + AI[:1]
    Shell_run_out_flipped = Shell_run_out[1:] + Shell_run_out[:1]

    # Number of variables
    num_vars = len(AI_flipped)

    # Create a 2D array for the flipped data
    values = [AI_flipped, Shell_run_out_flipped]

    # Create angles for the radar chart
    angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()

    values = [np.concatenate((value, [value[0]])) for value in values]
    angles += angles[:1]

    # Draw radar chart
    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)
    plt.xticks(angles[:-1], [f' {i+1}' for i in range(num_vars)], color='grey', size=8)
    ax.set_ylim(-100, 80) 

# Remove y-axis tick labels
    ax.set_yticks([])
    ax.plot(angles, values[0], color='red', linewidth=2, label='Ref')

    ax.plot(angles, values[1], color='blue', linewidth=2, label='Actual')


    ax.set_title(title, size=12, y=1.1)
    ax.legend(loc='upper right', bbox_to_anchor=(0.1, 0.1))

# Function to create a PDF with data, logo, and radar chart
def create_pdf_with_charts(excel_file, output_pdf, company_name, equipment_name, feed_rate, date_of_measurement, logo_path):



    # Load Excel file
    try:
        excel_data = pd.ExcelFile(excel_file)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    # User input for sheet-category mapping
    categories = ["Inlet", "Outlet", "TDN", "TUP", "GGDN", "GGUP"]
    category_images = {category: f"static/{category}.jpg" for category in categories}

    
    sheet_category_mapping = {}
    for category in categories:
        user_input = request.form[category.lower()]  # Get the input from the form
        sheet_indices = [int(i.strip()) - 1 for i in user_input.split(',') if i.strip().isdigit()]
        sheet_category_mapping[category] = sheet_indices

        all_sheets = set(range(len(excel_data.sheet_names)))  # Assuming excel_data is defined
        assigned_sheets = set(sheet for indices in sheet_category_mapping.values() for sheet in indices)
        unassigned_sheets = all_sheets - assigned_sheets
    
    notd_image = "static/NOTD.jpg"
    # Initialize PDF
    pdf = FPDF()
    class PDF(FPDF):
     def footer(self):
        # Set position for the footer
          self.set_y(-15)  # Position at 15 mm from the bottom
          self.set_font('Arial', 'I', 10)  # Set font for the footer
          self.set_text_color(255, 0, 0)  # Set text color to red
          self.cell(0, 10, 'Allan Smith Engineering Pvt. Ltd.', 0, 0, 'R')  # Centered footer

# Create an instance of the PDF class
    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)
        
            # Add Summary sheet to PDF
    try:
        summary_data = pd.read_excel(excel_file, sheet_name='Summary', header=4)  # Start from the 5th row
       # summary_data = summary_data.loc[:, summary_data.columns.difference(['X', 'Y'], sort=False)] 
        summary_data = summary_data.drop(columns=['X', 'Y'], errors='ignore')
        #summary_data = summary_data.loc[:, summary_data.columns.difference(['Bludge In', 'Bludge Out'], sort=False)]  # Skip columns X and Y
        





        pdf.add_page()
        pdf.set_font("Arial", style="B", size=16)
        pdf.cell(0, 10, "Summary", ln=True, align='C')
        pdf.ln(5)  # Add space after title

        # Add company info
        pdf.set_font("Arial",style="B", size=10)
        pdf.set_xy(35, 20)
        pdf.cell(0, 10, f"Company Name: {company_name}", ln=True)
        pdf.set_xy(35, 25)
        pdf.cell(0, 10, f"Equipment: {equipment_name}", ln=True)
        pdf.set_xy(35, 30)
        pdf.cell(0, 10, f"Feed Rate: {feed_rate}", ln=True)
        pdf.set_xy(35, 35)
        pdf.cell(0, 10, f"Date of Measurement: {date_of_measurement}", ln=True)

 
        pdf.ln(5) # Add space after company info

        # # Add logo
        if os.path.exists(logo_path):
           pdf.image(logo_path, x=10, y=21, w=20)  # Adjust logo position and size

        # Add summary table
        first_5_colsh = summary_data.iloc[:, :7]
        pdf.set_font("Arial", size=8)
        #col_widths = [20] * len(summary_data.columns)  # Adjust widths as needed
        headers = summary_data.columns.tolist()
        col_widthsl = [15, 30, 30, 20, 40,25,35,20,20]
        # Add column headers
        
        headers = first_5_colsh.columns.tolist()
        table_start_y = 50 # Adjust this value to move the table up or down as needed
        pdf.set_xy(10, table_start_y)
        table_data1 = first_5_colsh.fillna("").values.tolist()
        
        pdf.set_font("Arial", style='B', size=9)
        for i, header in enumerate(headers):
            pdf.cell(col_widthsl[i], 8, str(header), border=1, align='C')
        pdf.ln(8)
        pdf.set_font("Arial", size=9)

        # Add table rows
        for row in table_data1:
            for i, cell in enumerate(row):
                if headers[i] == "Position":  # Check if the current column is "Position"
                    pdf.cell(col_widthsl[i], 8, str(int(cell)), border=1, align='C')
                elif headers[i] == "Distance":  # Check if the current column is "Position"
                    pdf.cell(col_widthsl[i], 8, str(int(cell)), border=1, align='C')
                elif headers[i] == "Cumulative Distance":  # Check if the current column is "Position"
                    pdf.cell(col_widthsl[i], 8, str(int(cell)), border=1, align='C')
                elif isinstance(cell, (int, float)):
                    pdf.cell(col_widthsl[i], 8, f"{cell:.2f}", border=1, align='C')  # Limit to 2 decimal places
                else:
                    pdf.cell(col_widthsl[i], 8, str(cell), border=1, align='C')
            pdf.ln(8)

    except Exception as e:
        print(f"Error processing Summary sheet: {e}")
    
    eccentricity_values = summary_data['Eccentricity (mm)'].dropna().tolist()
    for idx, sheet_name in enumerate(excel_data.sheet_names):
        if sheet_name in ['Summary', 'Temp']:
            continue
    
        eccentricity_value = eccentricity_values[idx] if idx < len(eccentricity_values) else "N/A"

    # Create a line chart for Eccentricity and Runout
    try:
        runout_values = summary_data['Runout'].dropna()  # Drop NaN values
         

        # Create a line chart
        plt.figure(figsize=(10, 6))
        plt.plot(range(1, len(eccentricity_values) + 1), eccentricity_values, label='Eccentricity', marker='o')
        plt.plot(range(1, len(runout_values) + 1), runout_values, label='Runout', marker='o')
        # max_y_value = max(max(eccentricity_values), max(runout_values))
        # # for category, indices in sheet_category_mapping.items():
        # #  if category not in ['TDN', 'GGDN']:  
        # #   for idx in indices:
        # #     if category == 'TUP':
        # #         plt.fill_between([idx, idx + 1], 0, max_y_value, color="red", alpha=0.3)
        # #     else:
        # #         plt.fill_between([idx, idx + 1], 0, max_y_value, alpha=0.3 )

        for idx, value in enumerate(eccentricity_values):
         plt.text(idx + 1, value, f'{value:.1f}', fontsize=9, ha='center', va='bottom')

# Annotate Runout values
        for idx, value in enumerate(runout_values):
         plt.text(idx + 1, value, f'{value:.1f}', fontsize=9, ha='center', va='bottom')







# Draw rectangles for ranges specified in TDN and TUP
         #if 'TDN' in sheet_category_mapping:
          #for i in range(len(sheet_category_mapping['TDN']) - 1):
           #plt.fill_between([sheet_category_mapping['TDN'][i], sheet_category_mapping['TDN'][i + 1]], 0, max(eccentricity_values), alpha=0.3, color='orange', label='TDN Area')
 
        #if 'GGDN' in sheet_category_mapping:
         #for i in range(len(sheet_category_mapping['GGDN']) - 1):
          #plt.fill_between([sheet_category_mapping['GGDN'][i], sheet_category_mapping['GGDN'][i + 1]], 0, max(eccentricity_values), alpha=0.3, color='green', label='GGDN Area')
     

        plt.title('')
        plt.xlabel('Section')
        plt.ylabel('Eccentricity')
        plt.legend()
        plt.grid()
        plt.xticks(np.arange(1, len(eccentricity_values) + 1, 1))
        plt.tight_layout()
        
  
        # Save the chart as an image
        eccentricity_chart_path = 'eccentricity_runout_chart.png'
        plt.savefig(eccentricity_chart_path)
        plt.close()  # Close the plot to free memory
        with PILImage.open(eccentricity_chart_path) as imgn:
         rotated_img = imgn.rotate(90, expand=True)  # Rotate the image
         rotated_img.save('rotated_eccentricity_runout_chart.png') 
        # Add a new page for the Eccentricity and Runout line chart



        pdf.add_page(orientation='L')
        pdf.set_font("Arial",style="B", size=16)
        pdf.cell(0, 5, " Shell eccentricity and Run out deformation Graph ", ln=True, align='C')
        pdf.set_font("Arial",style="B", size=10)
        pdf.set_xy(35, 15)
        pdf.cell(0, 10, f"Company Name: {company_name}", ln=True)
        pdf.set_xy(35, 20)
        pdf.cell(0, 10, f"Equipment: {equipment_name}", ln=True)
        pdf.set_xy(35, 25)
        pdf.cell(0, 10, f"Feed Rate: {feed_rate}", ln=True)
        pdf.set_xy(35, 30)
        pdf.cell(0, 10, f"Date of Measurement: {date_of_measurement}", ln=True)
        pdf.ln(10)  # Add some space
        if os.path.exists(logo_path):
            pdf.image(logo_path, x=10, y=18, w=20)  # Add some space

        # Add the chart image to the PDF
        pdf.image('eccentricity_runout_chart.png', x=10, y=50, w=250)  # Adjust x, y, w as needed

        # # Add company details again for the Eccentricity and Runout section
        # pdf.set_font("Arial",style="B", size=10)
        # pdf.set_xy(35, 10)
        # pdf.cell(0, 10, f"Company Name: {company_name}", ln=True)
        # pdf.set_xy(35, 15)
        # pdf.cell(0, 10, f"Equipment: {equipment_name}", ln=True)
        # pdf.set_xy(35, 20)
        # pdf.cell(0, 10, f"Feed Rate: {feed_rate}", ln=True)
        # pdf.set_xy(35, 25)
        # pdf.cell(0, 10, f"Date of Measurement: {date_of_measurement}", ln=True)
        # pdf.set_xy(130, 40)
        # pdf.cell(0, 10, f" ALLAN SMITH ENGINEERING PVT LTD")
      
        # pdf.ln(5)   # Add space after company info

        # # Add logo again
        # if os.path.exists(logo_path):
        #     pdf.image(logo_path, x=10, y=10, w=20)  # Adjust logo position and size

    except Exception as e:
        print(f"Error creating Eccentricity and Runout line chart: {e}")

    # Add Temp sheet to PDF
    try:
        temp_data = pd.read_excel(excel_file, sheet_name='Temp', header=4)
        pdf.add_page()
        
        pdf.set_xy(10, 3)
        pdf.set_font("Arial", style="B", size=16)
        pdf.cell(0, 10, "Temperature Data", ln=True, align='C')
        pdf.ln(5)  # Add space after title

        # Add company info
        
        pdf.set_font("Arial",style="B", size=10)
        pdf.set_xy(35, 10)
        pdf.cell(0, 10, f"Company Name: {company_name}", ln=True)
        pdf.set_xy(35, 15)
        pdf.cell(0, 10, f"Equipment: {equipment_name}", ln=True)
        pdf.set_xy(35, 20)
        pdf.cell(0, 10, f"Feed Rate: {feed_rate}", ln=True)
        pdf.set_xy(35, 25)
        pdf.cell(0, 10, f"Date of Measurement: {date_of_measurement}", ln=True)

        pdf.ln(5)  # Add space after company info

        # Add logo
        if os.path.exists(logo_path):
            pdf.image(logo_path, x=10, y=10, w=20)  # Adjust logo position and size

        # Add temp table
        pdf.set_font("Arial", size=8)
        col_widths = [20] * len(temp_data.columns)  # Adjust widths as needed
        headers = temp_data.columns.tolist()
        
        table_start_y = 40 # Adjust this value to move the table up or down as needed
        pdf.set_xy(10, table_start_y)
        # Add column headers
        pdf.set_font("Arial", style='B', size=9)
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 6, str(header), border=1, align='C')
        pdf.ln(6)

        pdf.set_font("Arial", size=8)

        # Add table rows
        for row in temp_data.values:
            for i, cell in enumerate(row):
                if headers[i] == "Position":  # Check if the current column is "Position"
                    pdf.cell(col_widths[i], 6, str(int(cell)), border=1, align='C')
                elif isinstance(cell, (int, float)):
                    pdf.cell(col_widths[i], 6, f"{cell:.2f}", border=1, align='C')  # Limit to 2 decimal places
                else:
                    pdf.cell(col_widths[i], 6, str(cell), border=1, align='C')
            pdf.ln(6)

    except Exception as e:
        print(f"Error processing Temp sheet: {e}")

    # Create a line chart for the Temp data
    try:
        min_values = temp_data['Min'].dropna()  # Drop NaN values
        max_values = temp_data['Max'].dropna()  # Drop NaN values
        avg_values = temp_data['AVG'].dropna()  # Drop NaN values
        diff_values = temp_data['Diff'].dropna()  # Drop NaN values

        # Create a line chart
        plt.figure(figsize=(10, 6))
        plt.plot(range(1, len(min_values) + 1), min_values, label='Min', marker='o', color='blue')    # Change color to blue
        plt.plot(range(1, len(max_values) + 1), max_values, label='Max', marker='o', color='red')     # Change color to red
        plt.plot(range(1, len(avg_values) + 1), avg_values, label='Avg', marker='o', color='green')   # Change color to green
        plt.plot(range(1, len(diff_values) + 1), diff_values, label='Diff', marker='o', color='orange')
        for idx, value in enumerate(min_values):
         plt.text(idx + 1, value, f'{value:.1f}', fontsize=9, ha='center', va='bottom')

# Annotate Runout values
        for idx, value in enumerate(max_values):
         plt.text(idx + 1, value, f'{value:.1f}', fontsize=9, ha='center', va='bottom')
        for idx, value in enumerate(avg_values):
         plt.text(idx + 1, value, f'{value:.1f}', fontsize=9, ha='center', va='bottom')
        for idx, value in enumerate(diff_values):
         plt.text(idx + 1, value, f'{value:.1f}', fontsize=9, ha='center', va='bottom')
          
        plt.title('')
        plt.xlabel('Measured Position')
        plt.ylabel('Shell temperature')
        plt.legend()
        plt.grid()
        plt.xticks(np.arange(1, len(min_values) + 1, 1))
        plt.tight_layout()
         
        # Save the chart as an image
        chart_image_path = 'line_chart.png'
        plt.savefig(chart_image_path)
        plt.close()  # Close the plot to free memory

        # Add a new page for the line chart
        pdf.add_page(orientation='L')
        pdf.set_font("Arial",style="B", size=16)
        pdf.cell(0, 5, " Shell temperature profile Graph", ln=True, align='C')
        pdf.set_font("Arial",style="B", size=10)
        pdf.set_xy(35, 15)
        pdf.cell(0, 10, f"Company Name: {company_name}", ln=True)
        pdf.set_xy(35, 20)
        pdf.cell(0, 10, f"Equipment: {equipment_name}", ln=True)
        pdf.set_xy(35, 25)
        pdf.cell(0, 10, f"Feed Rate: {feed_rate}", ln=True)
        pdf.set_xy(35, 30)
        pdf.cell(0, 10, f"Date of Measurement: {date_of_measurement}", ln=True)
        pdf.ln(10)  # Add some space
        if os.path.exists(logo_path):
            pdf.image(logo_path, x=10, y=18, w=20)
        # Add the chart image to the PDF
        pdf.image(chart_image_path, x=10, y=50, w=250)  # Adjust x, y, w as needed

    except Exception as e:
        print(f"Error creating line chart: {e}")



    for idx, sheet_name in enumerate(excel_data.sheet_names):
        if sheet_name in ['Summary', 'Temp']:
            continue  # Skip already processed sheets
        eccentricity_value = eccentricity_values[idx] if idx < len(eccentricity_values) else "N/A"
        try:
            sheet_data = pd.read_excel(excel_file, sheet_name=sheet_name, header=4)

            # Clean column names
            sheet_data.columns = sheet_data.columns.str.strip()
            print(f"Processing sheet: {sheet_name}")
            print("Columns found:", sheet_data.columns.tolist())  # Print column names

            if 'AI' in sheet_data.columns and 'Shell Run Out' in sheet_data.columns:
                AI_data = sheet_data['AI'].dropna().tolist()
                Shell_run_out_data = sheet_data['Shell Run Out'].dropna().tolist()

                # Prepare PDF page
                pdf.add_page()
                pdf.set_font("Arial",style="B", size=10)

                # Add company logo
                if os.path.exists(logo_path):
                    pdf.image(logo_path, x=10, y=10, w=20)# Adjust logo position and size
                else:
                    print(f"Logo not found at {logo_path}. Skipping logo.")

                pdf.set_xy(10, 3)
                
                pdf.set_font("Arial", style="B", size=16)
                pdf.cell(0, 10, "Radial Run out eccentricity Polar Graph", ln=True, align='C')
                pdf.set_font("Arial", style="B", size=10)
                pdf.set_xy(35, 10)
                pdf.cell(0, 10, f"Company Name: {company_name}", ln=True)
                pdf.set_xy(35, 15)
                pdf.cell(0, 10, f"Equipment: {equipment_name}", ln=True)
                pdf.set_xy(35, 20)
                pdf.cell(0, 10, f"Feed Rate: {feed_rate}", ln=True)
                pdf.set_xy(35, 25)
                pdf.cell(0, 10, f"Date of Measurement: {date_of_measurement}")
  
                pdf.ln(5) # Add a little space after the header
                match = re.search(r'\d+$', sheet_name)
                sheet_number = match.group() if match else ""  
                pdf.set_xy(35, 30)
                pdf.cell(0, 10, txt=f"MeasuredPosition: {sheet_number}", ln=True, align='L')

               # Add Eccentricity, Bludge Out, Bludge In, and Angle of Occurrence values for this radar chart
                Min_temp_values = temp_data['Min'].dropna().tolist()
                Max_temp_values = temp_data['Max'].dropna().tolist()
                AVG_temp_values = temp_data['AVG'].dropna().tolist()

                
                angle_of_occurrence_values = summary_data['Phase Angle'].dropna().tolist()
                bludge_in_values = summary_data['Bludge In'].dropna().tolist()
                bludge_out_values = summary_data['Bludge Out'].dropna().tolist()
                runout_values = summary_data['Runout'].dropna().tolist()
                
                runout_values = runout_values[idx] if idx < len(runout_values) else "N/A"
                angle_of_occurrence_value = angle_of_occurrence_values[idx] if idx < len(angle_of_occurrence_values) else "N/A"
                bludge_out_values = bludge_out_values[idx] if idx < len(bludge_out_values) else "N/A"
                bludge_in_values = bludge_in_values[idx] if idx < len(bludge_in_values) else "N/A"
                Min_temp_value = Min_temp_values[idx] if idx < len(Min_temp_values) else "N/A"
                Max_temp_value = Max_temp_values[idx] if idx < len(Max_temp_values) else "N/A"
                AVG_temp_value = AVG_temp_values[idx] if idx < len(AVG_temp_values) else "N/A"

                # Add values to PDF
                pdf.set_font("Arial", style="B", size=13)
                pdf.set_xy(140, 40)
                pdf.cell(0, 10, f"Section Parameters:")
                pdf.set_font("Arial", size=13)
                pdf.set_xy(140, 48)
                pdf.cell(0, 10, f"Runout = {runout_values:.2f} mm" if isinstance(runout_values, (int, float)) else f"Runout = {runout_values} mm", ln=True,align='R')
                pdf.set_xy(140, 56)
                pdf.cell(0, 10, f"Eccentricity = {eccentricity_value:.2f} mm" if isinstance(eccentricity_value, (int, float)) else f"Eccentricity = {eccentricity_value} mm", ln=True,align='R')

                pdf.set_xy(140, 64)
                pdf.cell(0, 10, f"Angle of Occurrence = {angle_of_occurrence_value:.2f}°" if isinstance(angle_of_occurrence_value, (int, float)) else f"Angle of Occurrence = {angle_of_occurrence_value}°", ln=True ,align='R')
                pdf.set_xy(140, 79)
                pdf.set_font("Arial", style="B", size=13)
                pdf.cell(0, 10, f"Local Deformation:")
                pdf.set_font("Arial",  size=13)
                pdf.set_xy(140, 87)
                pdf.cell(0, 10, f"Bludge In   = {bludge_out_values:.2f} mm" if isinstance(bludge_out_values, (int, float)) else f"Bludge In  = {bludge_out_values} mm", ln=True,align='R')
                pdf.set_xy(140, 95)
                pdf.cell(0, 10, f"Bludge Out = {bludge_in_values:.2f} mm" if isinstance(bludge_in_values, (int, float)) else f"Bludge Out = {bludge_in_values} mm", ln=True,align='R')
                pdf.set_xy(140, 110)
                pdf.set_font("Arial", style="B", size=13)
                pdf.cell(0, 10, f"Temperature:")
                pdf.set_font("Arial",  size=13)
                pdf.set_xy(140, 118)
                pdf.cell(0, 10, f"MIN Temp = {Min_temp_value:.2f}°" if isinstance(Min_temp_value, (int, float)) else f"MIN Temp = {Min_temp_value}°", ln=True,align='R')
                pdf.set_xy(140, 126)
                pdf.cell(0, 10, f"MAX Temp = {Max_temp_value:.2f}°" if isinstance(Max_temp_value, (int, float)) else f"MAX Temp = {Max_temp_value}°", ln=True,align='R')
                pdf.set_xy(140, 134)
                pdf.cell(0, 10, f"AVG Temp = {AVG_temp_value:.2f}°" if isinstance(AVG_temp_value, (int, float)) else f"AVG Temp = {AVG_temp_value}°", ln=True,align='R')

  
  
                
    
                # Save radar chart to an image
                fig, ax = plt.subplots(figsize=(4, 4), subplot_kw=dict(polar=True))
                create_radar_chart(ax, AI_data, Shell_run_out_data, title=f"Shell profile eccentricity Diagram")
                radar_chart_path = f"{sheet_name}_radar_chart.png"
                plt.savefig(radar_chart_path, bbox_inches='tight', dpi=300)
                plt.close()

                first_5_cols = sheet_data.iloc[:, :5]
                pdf.set_font("Arial", size=8)

                col_widths = [15, 25, 25, 25, 24]
                # Add column headers
                table_start_y = 40 # Adjust this value to move the table up or down as needed
                pdf.set_xy(10, table_start_y)
                headers = first_5_cols.columns.tolist()

                pdf.set_font("Arial", style='B', size=9)
                for i, header in enumerate(headers):
                    pdf.cell(col_widths[i], 6, str(header), border=1, align='C')
                pdf.ln(6)
                # Add table rows

                pdf.set_font("Arial", size=8)
                table_data = first_5_cols.fillna("").values.tolist()
                for row in table_data:
                    for i, cell in enumerate(row[:5]):
                        if headers[i] == "Position":  # Check if the current column is "Position"
                            pdf.cell(col_widths[i], 6, str(int(cell)), border=1, align='C')
                        elif isinstance(cell, (int, float)):
                            pdf.cell(col_widths[i], 6, f"{cell:.2f}", border=1, align='C')  # Limit to 2 decimal places
                        else:
                            pdf.cell(col_widths[i], 6, str(cell), border=1, align='C')
                    pdf.ln(6)
                    
                # Add radar chart
                pdf.image(radar_chart_path, x=126, y=150, w=80)
                os.remove(radar_chart_path)  # Cleanup image
                # Add category-specific image if sheet matches user-specified category
                image_added = False
                for category, indices in sheet_category_mapping.items():
                    if idx in indices:
                        category_image_path = category_images.get(category, None)
                        if category_image_path and os.path.exists(category_image_path):
                            pdf.image(category_image_path, x=150, y=230, w=20)
                            image_added = True
                        else:
                            print(f"Category image for {category} not found or not defined.")
                        break

                # Add "NOTD" image if no other category matches
                if not image_added and idx in unassigned_sheets:
                    if os.path.exists(notd_image):
                        pdf.image(notd_image, x=150, y=230, w=20)
                    else:
                        print(f"NOTD image not found at {notd_image}. Skipping image.")

            else:
                print(f"Sheet {sheet_name} skipped: Required columns not found.")
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {e}")

    # Save PDF
    pdf.output(output_pdf)
    print(f"PDF saved at {output_pdf}")



def gear_radial_runout(excel_file, output_pdf, company_name, equipment_name, feed_rate, date_of_measurement, logo_path):
    # Load Excel file
    
    try:
        excel_data = pd.ExcelFile(excel_file)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    # User input for sheet-category mapping
    categories = ["Inlet", "Outlet", "TDN", "TUP", "GGDN", "GGUP"]
    category_images = {category: f"static/{category}.jpg" for category in categories}

    sheet_category_mapping = {}
    for category in categories:
        user_input = request.form[category.lower()]  # Get the input from the form
        sheet_indices = [int(i.strip()) - 1 for i in user_input.split(',') if i.strip().isdigit()]
        sheet_category_mapping[category] = sheet_indices

    all_sheets = set(range(len(excel_data.sheet_names)))  # Assuming excel_data is defined
    assigned_sheets = set(sheet for indices in sheet_category_mapping.values() for sheet in indices)
    unassigned_sheets = all_sheets - assigned_sheets

    notd_image = "static/NOTD.jpg"

    pdf = FPDF()
    
    summary_data = pd.read_excel(excel_file, sheet_name='Summary', header=4) 
    temp_data = pd.read_excel(excel_file, sheet_name='Temp', header=4)
    # Iterate through each sheet in the Excel file
    eccentricity_values = summary_data['Eccentricity (mm)'].dropna().tolist()
    for idx, sheet_name in enumerate(excel_data.sheet_names):
        if sheet_name in ['Summary', 'Temp']:
            continue  # Skip already processed sheets
        eccentricity_value = eccentricity_values[idx] if idx < len(eccentricity_values) else "N/A"
        try:
            sheet_data = pd.read_excel(excel_file, sheet_name=sheet_name, header=4)

            # Clean column names
            sheet_data.columns = sheet_data.columns.str.strip()
            print(f"Processing sheet: {sheet_name}")
            print("Columns found:", sheet_data.columns.tolist())  # Print column names

            if 'AI' in sheet_data.columns and 'Shell Run Out' in sheet_data.columns:
                AI_data = sheet_data['AI'].dropna().tolist()
                Shell_run_out_data = sheet_data['Shell Run Out'].dropna().tolist()

                # Prepare PDF page
                pdf.add_page()
                pdf.set_font("Arial",style="B", size=10)

                # Add company logo
                if os.path.exists(logo_path):
                    pdf.image(logo_path, x=10, y=10, w=20)# Adjust logo position and size
                else:
                    print(f"Logo not found at {logo_path}. Skipping logo.")

                pdf.set_xy(10, 3)
                
                pdf.set_font("Arial", style="B", size=16)
                pdf.cell(0, 10, "Radial Run out eccentricity Polar Graph", ln=True, align='C')
                pdf.set_font("Arial", style="B", size=10)
                pdf.set_xy(35, 10)
                pdf.cell(0, 10, f"Company Name         : {company_name}", ln=True)
                pdf.set_xy(35, 15)
                pdf.cell(0, 10, f"Equipment                 : {equipment_name}", ln=True)
                pdf.set_xy(35, 20)
                pdf.cell(0, 10, f"Feed Rate                   : {feed_rate}", ln=True)
                pdf.set_xy(35, 25)
                pdf.cell(0, 10, f"Date of Measurement: {date_of_measurement}")
  
                pdf.ln(5) # Add a little space after the header
                match = re.search(r'\d+$', sheet_name)
                sheet_number = match.group() if match else ""  
                pdf.set_xy(35, 30)
                pdf.cell(0, 10, txt=f"MeasuredPosition: {sheet_number}", ln=True, align='L')

               # Add Eccentricity, Bludge Out, Bludge In, and Angle of Occurrence values for this radar chart
                Min_temp_values = temp_data['Min'].dropna().tolist()
                Max_temp_values = temp_data['Max'].dropna().tolist()
                AVG_temp_values = temp_data['AVG'].dropna().tolist()

                
                angle_of_occurrence_values = summary_data['Phase Angle'].dropna().tolist()
                bludge_in_values = summary_data['Bludge In'].dropna().tolist()
                bludge_out_values = summary_data['Bludge Out'].dropna().tolist()
                runout_values = summary_data['Runout'].dropna().tolist()
                
                runout_values = runout_values[idx] if idx < len(runout_values) else "N/A"
                angle_of_occurrence_value = angle_of_occurrence_values[idx] if idx < len(angle_of_occurrence_values) else "N/A"
                bludge_out_values = bludge_out_values[idx] if idx < len(bludge_out_values) else "N/A"
                bludge_in_values = bludge_in_values[idx] if idx < len(bludge_in_values) else "N/A"
                Min_temp_value = Min_temp_values[idx] if idx < len(Min_temp_values) else "N/A"
                Max_temp_value = Max_temp_values[idx] if idx < len(Max_temp_values) else "N/A"
                AVG_temp_value = AVG_temp_values[idx] if idx < len(AVG_temp_values) else "N/A"

                # Add values to PDF
                pdf.set_font("Arial", style="B", size=13)
                pdf.set_xy(140, 40)
                pdf.cell(0, 10, f"Section Parameters:")
                pdf.set_font("Arial", size=13)
                pdf.set_xy(140, 48)
                pdf.cell(0, 10, f"Runout = {runout_values:.2f} mm" if isinstance(runout_values, (int, float)) else f"Runout = {runout_values} mm", ln=True,align='R')
                pdf.set_xy(140, 56)
                pdf.cell(0, 10, f"Eccentricity = {eccentricity_value:.2f} mm" if isinstance(eccentricity_value, (int, float)) else f"Eccentricity = {eccentricity_value} mm", ln=True,align='R')

                pdf.set_xy(140, 64)
                pdf.cell(0, 10, f"Angle of Occurrence = {angle_of_occurrence_value:.2f}°" if isinstance(angle_of_occurrence_value, (int, float)) else f"Angle of Occurrence = {angle_of_occurrence_value}°", ln=True ,align='R')
                pdf.set_xy(140, 79)
                pdf.set_font("Arial", style="B", size=13)
                pdf.cell(0, 10, f"Local Deformation:")
                pdf.set_font("Arial",  size=13)
                pdf.set_xy(140, 87)
                pdf.cell(0, 10, f"Bludge In   = {bludge_out_values:.2f} mm" if isinstance(bludge_out_values, (int, float)) else f"Bludge In  = {bludge_out_values} mm", ln=True,align='R')
                pdf.set_xy(140, 95)
                pdf.cell(0, 10, f"Bludge Out = {bludge_in_values:.2f} mm" if isinstance(bludge_in_values, (int, float)) else f"Bludge Out = {bludge_in_values} mm", ln=True,align='R')
                pdf.set_xy(140, 110)
                pdf.set_font("Arial", style="B", size=13)
                pdf.cell(0, 10, f"Temperature:")
                pdf.set_font("Arial",  size=13)
                pdf.set_xy(140, 118)
                pdf.cell(0, 10, f"MIN Temp = {Min_temp_value:.2f}°" if isinstance(Min_temp_value, (int, float)) else f"MIN Temp = {Min_temp_value}°", ln=True,align='R')
                pdf.set_xy(140, 126)
                pdf.cell(0, 10, f"MAX Temp = {Max_temp_value:.2f}°" if isinstance(Max_temp_value, (int, float)) else f"MAX Temp = {Max_temp_value}°", ln=True,align='R')
                pdf.set_xy(140, 134)
                pdf.cell(0, 10, f"AVG Temp = {AVG_temp_value:.2f}°" if isinstance(AVG_temp_value, (int, float)) else f"AVG Temp = {AVG_temp_value}°", ln=True,align='R')

  
  
                
    
                # Save radar chart to an image
                fig, ax = plt.subplots(figsize=(4, 4), subplot_kw=dict(polar=True))
                create_radar_chart(ax, AI_data, Shell_run_out_data, title=f"Shell profile eccentricity Diagram")
                radar_chart_path = f"{sheet_name}_radar_chart.png"
                plt.savefig(radar_chart_path, bbox_inches='tight', dpi=300)
                plt.close()

                first_5_cols = sheet_data.iloc[:, :5]
                pdf.set_font("Arial", size=8)

                col_widths = [15, 25, 25, 25, 24]
                # Add column headers
                table_start_y = 40 # Adjust this value to move the table up or down as needed
                pdf.set_xy(10, table_start_y)
                headers = first_5_cols.columns.tolist()

                pdf.set_font("Arial", style='B', size=9)
                for i, header in enumerate(headers):
                    pdf.cell(col_widths[i], 6, str(header), border=1, align='C')
                pdf.ln(6)
                # Add table rows

                pdf.set_font("Arial", size=8)
                table_data = first_5_cols.fillna("").values.tolist()
                for row in table_data:
                    for i, cell in enumerate(row[:5]):
                        if headers[i] == "Position":  # Check if the current column is "Position"
                            pdf.cell(col_widths[i], 6, str(int(cell)), border=1, align='C')
                        elif isinstance(cell, (int, float)):
                            pdf.cell(col_widths[i], 6, f"{cell:.2f}", border=1, align='C')  # Limit to 2 decimal places
                        else:
                            pdf.cell(col_widths[i], 6, str(cell), border=1, align='C')
                    pdf.ln(6)
                    
                # Add radar chart
                pdf.image(radar_chart_path, x=126, y=150, w=80)
                os.remove(radar_chart_path)  # Cleanup image
                # Add category-specific image if sheet matches user-specified category
                image_added = False
                for category, indices in sheet_category_mapping.items():
                    if idx in indices:
                        category_image_path = category_images.get(category, None)
                        if category_image_path and os.path.exists(category_image_path):
                            pdf.image(category_image_path, x=150, y=230, w=20)
                            image_added = True
                        else:
                            print(f"Category image for {category} not found or not defined.")
                        break

                # Add "NOTD" image if no other category matches
                if not image_added and idx in unassigned_sheets:
                    if os.path.exists(notd_image):
                        pdf.image(notd_image, x=150, y=230, w=20)
                    else:
                        print(f"NOTD image not found at {notd_image}. Skipping image.")

            else:
                print(f"Sheet {sheet_name} skipped: Required columns not found.")
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {e}")

    # Save PDF
    pdf.output(output_pdf)
    print(f"PDF saved at {output_pdf}")


@app.route('/')

def home():
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            return redirect(url_for('upload'))
        else:
            flash("Invalid username or password. Try again.", "danger")
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        new_user = User(username=username, password=hashed_password)
        db.session.add(new_user)
        db.session.commit()
        flash("Registration successful! Please login.", "success")
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        username = request.form['username']
        new_password = request.form['new_password']
        user = User.query.filter_by(username=username).first()
        if user:
            user.password = generate_password_hash(new_password, method='pbkdf2:sha256')
            db.session.commit()
            flash("Password reset successful. Please login.", "success")
            return redirect(url_for('login'))
        else:
            flash("Username not found.", "danger")
    return render_template('forgot_password.html')

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if 'user_id' not in session:
        return redirect(url_for('login'))  # Redirect to login if user is not authenticated

    if request.method == 'POST':
        if 'file' not in request.files:
            flash("No file part", "danger")
            return redirect(url_for('upload'))  # Redirect back to the upload page
        
        file = request.files['file']
        if file.filename == '':
            flash("No selected file", "danger")
            return redirect(url_for('upload'))  # Redirect back to the upload page

        # Save the uploaded file
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        # Get other form data
        selected_module = request.form.get("module")
        company_name = request.form['company_name']
        equipment_name = request.form['equipment_name']
        feed_rate = request.form['feed_rate']
        date_of_measurement = request.form['date_of_measurement']
        logo_path = 'static/companylogo.jpg'  # Adjust as necessary

        # Get the number of positions from the form
        user_inp = request.form['position']
        user_inp_int = int(user_inp) 
        angle_increment = 360 / user_inp_int

        # Load Excel file and process data
        read_sheet_name = 0  # Adjust if needed (default first sheet)
        data = pd.read_excel(file_path, sheet_name=read_sheet_name)

        # Extract rows where 'CHAIRPAD NO' is present and numeric data follows
        filtered_data = data[data['CHAIRPAD NO'].apply(lambda x: str(x).isnumeric())]

        # Relabel columns to A, B, C, etc.
        filtered_data.columns = [chr(65 + i) for i in range(len(filtered_data.columns))]
        
        position = list(range(1, user_inp_int + 1)) + [1]  # Use user input for positions
        measurement = [i * angle_increment for i in range(user_inp_int)] + [360]  # Adjust measurement based on user input

        # Initialize the summary data
        summary_data = []

        # Locate Distance and Cumulative Distance values
        distance_row = data.iloc[68, 1:].values
        cumulative_distance_row = data.iloc[69, 1:].values
        Temp_data = []
        Diff_temp = data.iloc[70, 1:].values - data.iloc[71, 1:].values
        Min_temp = data.iloc[70, 1:].values
        Max_temp = data.iloc[71, 1:].values
        AVG_temp = data.iloc[72, 1:].values

        # Create separate sheets for each column starting from B
        output_file = 'processed_data_with_summary.xlsx'
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for i, col in enumerate(filtered_data.columns[1:]):  # Start from column B
                sheet_name = f"Sheet_{i + 1}"

                # Create the sheet data
                data_measured = filtered_data[col].tolist() + [filtered_data[col].iloc[0]]
                if len(data_measured) < len(position):
                    data_measured += [np.nan] * (len(position) - len(data_measured))

                # Ensure data_measured contains numeric values, coercing errors to NaN
                data_measured = pd.to_numeric(data_measured, errors='coerce')

                # Calculate Shell Run Out
                shell_run_out = [np.nanmax(data_measured) - value if not np.isnan(value) else 0 for value in data_measured]

                # Ensure all lists are of the same length
                max_length = max(len(position), len(measurement), len(data_measured), len(shell_run_out))

                # Extend lists to match the maximum length
                position.extend([np.nan] * (max_length - len(position)))
                measurement.extend([np.nan] * (max_length - len(measurement)))
                data_measured = data_measured.tolist() + [np.nan] * (max_length - len(data_measured))
                shell_run_out.extend([np.nan] * (max_length - len(shell_run_out)))

                # Create the DataFrame
                sheet_data = pd.DataFrame({
                    'Position': position[:max_length],
                    'Measurement': measurement[:max_length],
                    'Data Measured': data_measured[:max_length],
                    'Distortion': data_measured[:max_length],
                    'Shell Run Out': shell_run_out[:max_length],
                })

                sheet_data['AA'] = sheet_data['Measurement'] / 180 * 3.14
                sheet_data['AB'] = np.cos(sheet_data['AA']) * sheet_data['Shell Run Out']
                sheet_data['AC'] = np.sin(sheet_data['AA']) * sheet_data['Shell Run Out']

                SUM_AB = sheet_data['AB'][:-1].sum()
                SUM_AC = sheet_data['AC'][:-1].sum()

                XX = 2 / user_inp_int * SUM_AB
                YY = 2 / user_inp_int * SUM_AC
                ZZ = np.sqrt(XX ** 2 + YY ** 2)

                Angle_of_Occurrence = np.arccos(XX / ZZ) * 180 / 3.14 if ZZ != 0 else 0
                if YY < 0:
                    Angle_of_Occurrence = 360 - Angle_of_Occurrence

                sheet_data['AD'] = (Angle_of_Occurrence - sheet_data['Measurement']) / 180 * 3.14
                sheet_data['AE'] = np.cos(sheet_data['AD'])  # Column name corrected to 'AE'
                sheet_data['AF'] = ZZ * sheet_data['AE']    # Column name corrected to 'AE'

                sheet_data['AG'] = sheet_data['Shell Run Out'] - sheet_data['AF']
                AVG_AG = sheet_data['AG'][:-1].mean()
                sheet_data['AH'] = sheet_data['AG'] - AVG_AG  # Calculate Distortion (AH)
                sheet_data['Distortion'] = sheet_data['AH']
                sheet_data['AI'] = sheet_data['AF'] + AVG_AG
                distance_value = distance_row[i]  # Access distance value using index i
                cumulative_distance_value = cumulative_distance_row[i]  # Access cumulative distance value using index i
                Diff_temp_value = Diff_temp[i]
                Min_temp_value = Min_temp[i]
                Max_temp_value = Max_temp[i]
                AVG_temp_value = AVG_temp[i]

                # Collect summary data
                max_runout = np.nanmax(shell_run_out)

                summary_data.append({
                    'Position': i + 1,
                    'X': XX,
                    'Y': YY,
                    'Eccentricity (mm)': ZZ,
                    'Phase Angle': Angle_of_Occurrence,
                    'Runout': max_runout,
                    'Local Shell Deformation': AVG_AG,
                    'Distance': distance_row[i] if i < len(distance_row) else np.nan,
                    'Cumulative Distance': cumulative_distance_row[i] if i < len(cumulative_distance_row) else np.nan,
                })
                Temp_data.append({
                    'Position': i + 1,
                    'Diff': Diff_temp[i] if i < len(Diff_temp) else np.nan,
                    'Min': Min_temp[i] if i < len(Min_temp) else np.nan,
                    'Max': Max_temp[i] if i < len(Max_temp) else np.nan,
                    'AVG': AVG_temp[i] if i < len(AVG_temp) else np.nan,
                })

                # Write sheet data
                sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)

            # Create the summary sheet
            summary_df = pd.DataFrame(summary_data)
            Temp_df = pd.DataFrame(Temp_data)
            Temp_df.to_excel(writer, sheet_name='Temp', index=False)

        # Add Bludge In and Bludge Out columns
        bludge_in = []
        bludge_out = []

        # Traverse through each sheet to find max and min distortion
        for i in range(len(filtered_data.columns) - 1):  # Adjust for the number of sheets
            sheet_name = f"Sheet_{i + 1}"
            sheet_data = pd.read_excel(output_file, sheet_name=sheet_name)

            # Find max and min distortion
            max_distortion = sheet_data['Distortion'].max()  # Assuming 'Distortion' is the correct column
            min_distortion = sheet_data['Distortion'].min()

            bludge_out.append(max_distortion)
            bludge_in.append(min_distortion)

        # Add the new columns to the summary DataFrame
        summary_df['Bludge In'] = bludge_in
        summary_df['Bludge Out'] = bludge_out

        # Write the updated summary DataFrame to the summary sheet
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Reload workbook to add logo and inputs
        workbook = load_workbook(output_file)

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            try:
                img = Image(logo_path)
                img.width = 100
                img.height = 50
                worksheet.add_image(img, 'A1')
            except FileNotFoundError:
                print(f"Logo not found. Skipping for {sheet_name}.")

            # Add inputs to the top rows with reduced cell sizes
            worksheet.insert_rows(1, amount=4)
            worksheet.cell(row=1, column=3, value="Plant").font = Font(size=8)
            worksheet.cell(row=1, column=4, value=company_name).font = Font(size=8)
            worksheet.cell(row=2, column=3, value="Equipment").font = Font(size=8)
            worksheet.cell(row=2, column=4, value=equipment_name).font = Font(size=8)
            worksheet.cell(row=3, column=3, value="Feed Rate").font = Font(size=8)
            worksheet.cell(row=3, column=4, value=feed_rate).font = Font(size=8)
            worksheet.cell(row=4, column=3, value="Date of Measurement").font = Font(size=8)
            worksheet.cell(row=4, column=4, value=date_of_measurement).font = Font(size=8)

            # Adjust column widths for better visibility
            for col in range(1, 5):
                worksheet.column_dimensions[get_column_letter(col)].width = 15

        workbook.save(output_file)
        print(f"Processed data with summary saved to: {output_file}")

        # Generate PDF report
        output_pdf = "Radar_Charts_Report_with_Logo.pdf"  # Desired output PDF file name
        module_functions = {
            "create_pdf_with_charts": create_pdf_with_charts,
            "gear_tyre_axial_runout": gear_tyre_axial_runout,
            "gear_radial_runout": gear_radial_runout,
            "support_roller_deflection": support_roller_deflection,
            "support_roller_raceway": support_roller_raceway,
        }

        # Check if the selected module exists in the dictionary
        function_to_call = module_functions.get(selected_module)

        if function_to_call:
            function_to_call(output_file, output_pdf, company_name, equipment_name, feed_rate, date_of_measurement, logo_path)
            return send_file(output_pdf, as_attachment=True)
        else:
            print("Invalid module selected!")  # Debugging output
            return "Invalid module selection", 400

    # If the request method is GET, render the upload form
    return render_template('upload.html')
    # If the request method is GET, render the upload form
        

if __name__ == '__main__':
 with app.app_context():
    db.create_all()
    app.run(debug=True)


    